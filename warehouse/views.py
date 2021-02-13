from django.shortcuts import render
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.request import Request
from rest_framework.response import Response
from .services import cargo_load, cargo_unload, cargo_transfer
from .models import WarehousePoint, Cargo, CargoUnloadTimeSheet, CargoLoadTimeSheet
from .serializers import CargoSerializer
from openpyxl import load_workbook
from datetime import datetime, date
import time


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def load_cargo_view(request: Request) -> Response:
    cargo_ser = CargoSerializer(request.data)
    if cargo_ser.is_valid():
        cargo = Cargo(**cargo_ser.validated_data)
        ware_point: WarehousePoint = WarehousePoint.objects.filter(type='VOID').first()
        data = cargo_load({"point": ware_point.point_num, "shelving": ware_point.shelving})
        if data.get('success'):
            cargo.save()
            return Response(data, 201)
        return Response(data, 400)
    else:
        return Response({"success": False, 'error': cargo_ser.errors})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def unload_cargo_view(request: Request) -> Response:
    if request.data.get('point') and request.data.get('shelving'):
        return Response({'success': False, 'error': 'неверные параметры'}, 422)
    data = cargo_unload(request.data)
    if data.get('success'):
        return Response(data, 201)
    return Response(data, 400)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def transfer_cargo_view(request: Request) -> Response:
    if request.data.get('point') and request.data.get('shelving'):
        return Response({'success': False, 'error': 'неверные параметры'}, 422)
    data = cargo_transfer(**request.data)
    if data.get('success'):
        return Response(data, 201)
    return Response(data, 400)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def load_cargo_file_view(request: Request) -> Response:
    file = request.FILES.get('file')
    if file:
        if '.xlsx' not in file.name:
            return Response({'success': False, 'error': 'Невереное расширение файла'}, 422)
        with open('media/tmp_factory/' + file.name, 'wb+') as f:
            for chunk in file.chunks():
                f.write(chunk)
        wb = load_workbook('media/tmp_factory/' + file.name, data_only=True)
        sheet = wb.active

        for i in range(2, sheet.max_row):
            name = sheet.cell(row=i, column=1).value
            model = sheet.cell(row=i, column=2).value
            factory_id = sheet.cell(row=i, column=3).value
            date_cargo = sheet.cell(row=i, column=4).value
            time_in = sheet.cell(row=i, column=5).value
            time_out = sheet.cell(row=i, column=6)
            producer = sheet.cell(row=i, column=7).value
            if len(name) < 1:
                break
            else:
                time_in = datetime.combine(date_cargo, time_in)
                time_out = datetime.combine(date_cargo, time_out)
                new_cargo = Cargo.objects.create(cargo_model=model, name=name,
                                                 factory_io_id=int(factory_id),
                                                 producer=producer,
                                                 place=None)
                CargoUnloadTimeSheet.objects.create(factory_io_id=int(factory_id),
                                                    date_unload=time_out)
                CargoLoadTimeSheet.objects.create(factory_io_id=int(factory_id),
                                                  date_load=time_in)

